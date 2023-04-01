import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import threading
import datetime

from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

from selenium.webdriver.common.keys import Keys

from selenium.common.exceptions import TimeoutException

chrome_options = Options()
chrome_options.headless = False
chrome_options.add_argument("start-maximized")
# options.add_experimental_option("detach", True)
chrome_options.add_argument("--no-sandbox")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_argument('--disable-blink-features=AutomationControlled')

from file_reader_class import CustomFileReader
f_c=CustomFileReader()
var_filename=f_c.path
search_var=['FULSTL','CLESTL','Short term']

def conn(var_lst_inp):
    # possible for loop
    var_sheet_name=search_var[0]
    global driver
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
    driver.get('https://citydev-portal.edinburgh.gov.uk/idoxpa-web/search.do?action=simple&searchType=Application')
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))).click()
        print('accepted cookies')
    except Exception as e:
        print('no cookie button!')

    WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.NAME, "searchCriteria.simpleSearchString"))).send_keys(
            var_lst_inp + Keys.RETURN)
    crawler(var_lst_inp)
    print('passed crawler in conn')


#Class returning from thread
class ThreadParseDate(threading.Thread):
    def __init__(self,var_xpath):
        threading.Thread.__init__(self)
        self.value='-'
        self.var_xp=var_xpath

    def run(self):
        try:
            var_wbdriver = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located(
                    (By.XPATH, self.var_xp))).text
            self.value = datetime.datetime.strptime(var_wbdriver, '%a %d %b %Y').strftime('%d/%m/%Y')
            print(var_wbdriver, 'threading link date',
                  datetime.datetime.strptime(var_wbdriver, '%a %d %b %Y').strftime('%d-%m-%Y'))
        except TimeoutException:
            self.value = '-'

class ThreadParseStatus(threading.Thread):
    def __init__(self,var_xpath):
        threading.Thread.__init__(self)
        self.value='-'
        self.var_xp=var_xpath

    def run(self):
        try:
            link_status = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located(
                    (By.XPATH, self.var_xp))).text
            self.value = link_status
            print(link_status, 'threading status decision note')
        except TimeoutException:
            self.value = '-'

class ThreadParseLinks(threading.Thread):
    def __init__(self,var_xpath):
        threading.Thread.__init__(self)
        self.value='-'
        self.var_xp=var_xpath

    def run(self):
            try:
                link_docs = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, self.var_xp)))
                self.value=link_docs.get_attribute('href')
                print(link_docs.get_attribute('href'), 'threading document links')
            except TimeoutException:
                self.value='-'


def link_date_appl():
    try:
        link_date_appl = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located(
                (By.XPATH, "//th[contains(text(),'Application Validated')]/following-sibling::td"))).text
        ret_val=(datetime.datetime.strptime(link_date_appl, '%a %d %b %Y').strftime('%d/%m/%Y'))
        print(link_date_appl, 'link date application',
              datetime.datetime.strptime(link_date_appl, '%a %d %b %Y').strftime('%d-%m-%Y'))
    except TimeoutException:
        ret_val=('-')

    return ret_val


def doc_link_parsing(var_inp_list:list,var_sheet_name):
    # Open a new window
    driver.execute_script("window.open('');")
    lst_date_appl = []
    lst_date_decision = []
    lst_decision_note=[]
    lst_doc_links = []
    for link in var_inp_list:
        # Switch to the new window and open new URL
        driver.switch_to.window(driver.window_handles[1])
        driver.get(link)

        print('start thread')
        thread_lnk_date_applied=ThreadParseDate("//th[contains(text(),'Application Validated')]/following-sibling::td")
        thread_lnk_date_decision=ThreadParseDate("//th[contains(text(),'Decision Issued Date')]/following-sibling::td")
        thread_status_decision = ThreadParseStatus("//th[contains(text(),'Decision')]/following-sibling::td")
        thread_doc_link=ThreadParseLinks("//p[@class='associateddocument']/a")
        print('not yet')
        thread_lnk_date_applied.start()
        thread_lnk_date_applied.join()
        thread_lnk_date_decision.start()
        thread_status_decision.start()
        thread_doc_link.start()

        thread_lnk_date_applied.join()
        thread_status_decision.join()
        thread_doc_link.join()

        ret_link_date_applied_value=thread_lnk_date_applied.value
        print('--date applied--',ret_link_date_applied_value)
        lst_date_appl.append(ret_link_date_applied_value)

        ret_link_date_decision_value=thread_lnk_date_decision.value
        print('--date decision--',ret_link_date_decision_value)
        lst_date_decision.append(ret_link_date_decision_value)

        ret_status_decision_value = thread_status_decision.value
        print('--status decision--', ret_status_decision_value)
        lst_decision_note.append(ret_status_decision_value)

        ret_link_doc_value = thread_doc_link.value
        print('--doc link--', ret_link_doc_value)
        lst_doc_links.append(ret_link_doc_value)

    # Close newly opened tab
    driver.execute_script("window.close('');")
    driver.switch_to.window(driver.window_handles[0])

    df_lst_date_app = pd.DataFrame(lst_date_appl)
    df_lst_date_dec_iss = pd.DataFrame(lst_date_decision)
    df_lst_link=pd.DataFrame(lst_doc_links)
    df_lst_decision_note=pd.DataFrame(lst_decision_note)

    write_docs_xl(df_lst_link,df_lst_decision_note,df_lst_date_dec_iss,df_lst_date_app,var_sheet_name)


def crawler(var_sheet_name):
    print('here in crawler')
    lst_title = []
    lst_link = []
    var_titles = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//li[@class='searchresult']/a")))
    for var_title in var_titles:
        lst_link.append(str(var_title.get_attribute('href')).strip())
        lst_title.append(str(var_title.get_attribute('text')).strip())

    lst_addr = []
    var_addresses = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//li[@class='searchresult']/p[@class='address']")))
    for var_address in var_addresses:
        lst_addr.append(str(var_address.get_attribute('innerHTML')).strip())

    lst_ref = []
    lst_recv = []
    lst_valid = []
    lst_status = []
    var_refs = WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located((By.XPATH, "//li[@class='searchresult']/p[@class='metaInfo']")))
    for var_ref in var_refs:
        lst_ref.append(str(var_ref.text).split('|')[0])
        lst_recv.append(datetime.datetime.strptime(str(var_ref.text).split('|')[1].split(':')[1].strip(), '%a %d %b %Y').strftime('%d-%m-%Y'))
        lst_valid.append(str(var_ref.text).split('|')[2])
        lst_status.append(str(var_ref.text).split('|')[3])

    # click next button
    # Pool making dataframes
    df_alldata = pd.DataFrame({'Title': lst_title, 'Links': lst_link, 'References': lst_ref, 'Address': lst_addr,
                               'Validated': lst_valid, 'Status': lst_status, 'Received': lst_recv})

    thread_click_next = threading.Thread(target=clicking_next_page, args=(var_sheet_name,))
    thread_second = threading.Thread(target=write_xl, args=('output.xlsx', df_alldata, var_sheet_name))
    thread__link=threading.Thread(target=doc_link_parsing,args=(lst_link,var_sheet_name))
    thread__link.start()
    thread__link.join()
    thread_click_next.start()
    thread_second.start()
    thread_click_next.join()
    thread_second.join()
    print('crawler finished here...')


def clicking_next_page(var_sheet_name):
    try:
        btn_next_to_click = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@class='next']")))
        btn_next_to_click.click()
        crawler(var_sheet_name)
        print('passed crawler in clicking next page')
    except TimeoutException:
        try:
            if len(search_var)>0:
                search_var.pop(0)
                conn(search_var[0])
            else:
                pass
        except TimeoutException or IndexError:
            driver.quit()
            print('finished')


def write_xl(var_path, df_alldata, var_sheet_name):
    var_path=var_filename
    var_sheet_name = var_sheet_name.replace(' ','')
    print(var_sheet_name)
    if os.path.isfile(var_path):
        with pd.ExcelWriter(var_path, engine='openpyxl', mode='r+', if_sheet_exists='overlay') as writer:
            book = load_workbook(var_path)
            current_sheet = book[var_sheet_name]
            Column_A = current_sheet['A']
            try:
                maxrow = max(c.row for c in Column_A if c.value is not None)
            except ValueError:
                maxrow = 0

            sheetname = var_sheet_name
            print(df_alldata)
            df_alldata.to_excel(writer, sheet_name=sheetname, startrow=maxrow, index=False, header=False)
    else:
        df_alldata.to_excel(var_path, sheet_name=var_sheet_name, index=False, header=True)
        print(df_alldata)


def write_docs_xl(var_ListDocLinks:pd.DataFrame,
                  var_ListStatus:pd.DataFrame,var_ListDecIssDate:pd.DataFrame,
                  var_ListAppValDate:pd.DataFrame,var_sheetname):
    var_sheetname= var_sheetname.replace(' ','')
    var_path=var_filename
    with pd.ExcelWriter(var_path, engine='openpyxl', mode='r+', if_sheet_exists='overlay') as writer:
        book = load_workbook(var_path)
        ws=book[var_sheetname]
        Column_H=ws['I']
        lst_max_cell_value_none=[]
        try:
            for cell in Column_H:
                if cell.value is not None:
                    lst_max_cell_value_none.append(cell.row)
            maxrow=(max(lst_max_cell_value_none))
            print(maxrow,'maximum row')
        except ValueError:
            maxrow = 1
        print(var_sheetname,maxrow,'writing docs')
        var_ListDocLinks.to_excel(writer,startrow=maxrow,startcol=10,sheet_name=var_sheetname,index=False,header=False)
        var_ListStatus.to_excel(writer, startrow=maxrow, startcol=7, sheet_name=var_sheetname, index=False, header=False)
        var_ListDecIssDate.to_excel(writer, startrow=maxrow, startcol=8, sheet_name=var_sheetname, index=False,
                                header=False)
        var_ListAppValDate.to_excel(writer, startrow=maxrow, startcol=9, sheet_name=var_sheetname, index=False,
                                    header=False)


def create_sheets():
    wb=Workbook()
    count=len(search_var)
    for n in range(count):
        wb.create_sheet(title=search_var[n].replace(' ',''))
        wb[search_var[n].replace(' ','')]['A1']='Title'
        wb[search_var[n].replace(' ','')]['B1'] = 'Links'
        wb[search_var[n].replace(' ','')]['C1'] = 'References'
        wb[search_var[n].replace(' ','')]['D1'] = 'Address'
        wb[search_var[n].replace(' ','')]['E1'] = 'Validated'
        wb[search_var[n].replace(' ','')]['F1'] = 'Status'
        wb[search_var[n].replace(' ','')]['G1'] = 'Received'
        wb[search_var[n].replace(' ', '')]['H1'] = 'Decision Note'
        wb[search_var[n].replace(' ', '')]['I1'] = 'Decision Issued Date'
        wb[search_var[n].replace(' ', '')]['J1'] = 'Application Validated'
        wb[search_var[n].replace(' ', '')]['K1'] = 'Document Link'
    del wb['Sheet']
    wb.save(var_filename)

def start_app():
    create_sheets()
    conn(search_var[0])

if __name__=='__main__':
    start_app()
